"""
lighting-ai/services/exporter/exporter.py

M6  DXF export  — professional Deckenrasterplan format
                  (title block, legend, dimension chains, ATTRIB blocks)
M7  Excel BOM   — 3-sheet workbook
M8  PDF/HTML    — customer documentation package
"""
from __future__ import annotations
import datetime
from collections import Counter, defaultdict
from pathlib import Path
from typing import Optional
import sys

import ezdxf
from ezdxf.enums import TextEntityAlignment
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from jinja2 import Template

sys.path.insert(0, str(Path(__file__).parent.parent.parent))
from config import EXPORTS_DIR
from services.placer.real_placer import PlacementResult, PlacedLuminaire, ZoneLightingReport
from services.classifier.room_classifier_real import ClassifiedPlan

# ── DXF layer / colour constants ─────────────────────────────────────────────
LUMI_LAYER  = "LUMINAIRES"
ZONE_LAYER  = "ZONES"
GRID_LAYER  = "CEILING-GRID"
DIM_LAYER   = "DIMENSIONS"
TITLE_LAYER = "TITLEBLOCK"
LEGEND_LAYER= "LEGEND"
ANNO_LAYER  = "ANNOTATIONS"

COLOR_A    = 6   # magenta    — K1 shelf interior standard
COLOR_AW   = 6   # magenta    — K1 exterior-wall shelf high-beam (same symbol as A)
COLOR_B    = 1   # red        — K4 supplement wide-angle
COLOR_C    = 4   # cyan       — K3 shelf edge / perimeter
COLOR_D    = 2   # yellow     — K2 checkout / service strong
COLOR_E    = 5   # blue       — K6 NEO85-SX window track
COLOR_W    = 3   # green      — honeycomb anti-glare cosmetics
COLOR_P    = 30  # orange     — K5 Plakate poster accent
COLOR_ZONE = 140 # light blue — zone outlines
COLOR_GRID = 8   # dark grey  — ceiling grid
COLOR_DIM  = 7   # white/black — dimensions

LUMI_COLORS = {'A': COLOR_A, 'AW': COLOR_AW, 'B': COLOR_B, 'C': COLOR_C,
               'D': COLOR_D, 'E': COLOR_E, 'W': COLOR_W, 'P': COLOR_P}

CUTOUT_MM   = 128   # DA (visible cutout diameter)
OUTER_MM    = 140   # AD (outer diameter)

# ─────────────────────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────────────────────

def _add_layer(doc, name, color, lw=25):
    if name not in doc.layers:
        doc.layers.add(name, dxfattribs={"color": color, "lineweight": lw})


def _txt(msp, text, pos, height, layer, color=7, align=TextEntityAlignment.LEFT):
    msp.add_text(text, dxfattribs={
        "layer": layer, "height": height, "color": color,
        "insert": pos,
    }).set_placement(pos, align=align)


# ─────────────────────────────────────────────────────────────────────────────
# Luminaire block with ATTDEF attributes
# ─────────────────────────────────────────────────────────────────────────────

def _ensure_lumi_block(doc, lumi_type: str, product_code: str,
                        cutout_mm: float, color: int) -> str:
    """
    Create (or reuse) a block for a luminaire type.
    The block contains:
      - Outer circle (cutout diameter)
      - Inner dot
      - Cross-hair
      - ATTDEFs for product code, type, wattage
    """
    # Sanitise block name: strip non-alphanumeric except _
    safe = (f"MIKA_{lumi_type}_" +
            "".join(c if c.isalnum() or c == '_' else '_'
                    for c in product_code))[:40]
    if safe in doc.blocks:
        return safe

    blk = doc.blocks.new(name=safe)
    r   = cutout_mm / 2

    # Type-specific geometry
    if lumi_type == 'A':
        # Circle + inner circle + orthogonal crosshair
        blk.add_circle((0,0), r,               dxfattribs={"layer":"0","color":color})
        blk.add_circle((0,0), r * 0.30,        dxfattribs={"layer":"0","color":color})
        blk.add_line((-r*.6,0),(r*.6,0),       dxfattribs={"layer":"0","color":color})
        blk.add_line((0,-r*.6),(0,r*.6),       dxfattribs={"layer":"0","color":color})
    elif lumi_type == 'B':
        # Square + center dot
        blk.add_lwpolyline([(-r,-r),(r,-r),(r,r),(-r,r)], close=True,
                           dxfattribs={"layer":"0","color":color})
        blk.add_circle((0,0), r * 0.20,        dxfattribs={"layer":"0","color":color})
    elif lumi_type == 'C':
        # Diamond + center dot
        blk.add_lwpolyline([(0,-r),(r,0),(0,r),(-r,0)], close=True,
                           dxfattribs={"layer":"0","color":color})
        blk.add_circle((0,0), r * 0.20,        dxfattribs={"layer":"0","color":color})
    elif lumi_type == 'D':
        # Equilateral triangle pointing down + center dot
        blk.add_lwpolyline([(-r*0.866, r*0.5),(r*0.866, r*0.5),(0,-r)], close=True,
                           dxfattribs={"layer":"0","color":color})
        blk.add_circle((0,0), r * 0.20,        dxfattribs={"layer":"0","color":color})
    elif lumi_type == 'E':
        # Circle + diagonal X cross (distinct from A's orthogonal +)
        blk.add_circle((0,0), r,               dxfattribs={"layer":"0","color":color})
        blk.add_line((-r*.6,-r*.6),(r*.6,r*.6),dxfattribs={"layer":"0","color":color})
        blk.add_line((-r*.6,r*.6),(r*.6,-r*.6),dxfattribs={"layer":"0","color":color})
    else:
        # Fallback: double circle
        blk.add_circle((0,0), r,               dxfattribs={"layer":"0","color":color})
        blk.add_circle((0,0), r * 0.35,        dxfattribs={"layer":"0","color":color})

    # ATTDEFs (invisible — carried in block for BOM extraction)
    attribs_base = {"layer":"0","height":r*0.4,"flags":ezdxf.const.ATTRIB_INVISIBLE}
    blk.add_attdef("TYPE",    insert=(0, r*1.2), text=lumi_type,    dxfattribs=attribs_base)
    blk.add_attdef("PRODUCT", insert=(0,-r*1.8), text=product_code, dxfattribs=attribs_base)

    return safe


# ─────────────────────────────────────────────────────────────────────────────
# Title block (Schriftfeld)
# ─────────────────────────────────────────────────────────────────────────────

def _draw_title_block(msp, x0: float, y0: float,
                      project_name: str, customer: str,
                      concept_id: str, scale: str,
                      total_lumi: int, total_w: float,
                      generated: str):
    """
    Draw a professional title block at (x0, y0).
    Width = 180mm, height = 60mm (in drawing units = mm).
    """
    W, H = 180_000, 60_000  # in drawing mm units (1 unit = 1mm)
    TH   = 3_000   # text height (3mm)
    LH   = 2_000   # label height (2mm)
    PAD  = 2_500

    def box(x, y, w, h):
        pts = [(x,y),(x+w,y),(x+w,y+h),(x,y+h)]
        msp.add_lwpolyline(pts, close=True,
                           dxfattribs={"layer": TITLE_LAYER, "color": 7})

    def label(text, x, y, h=LH, color=7):
        _txt(msp, text, (x, y), h, TITLE_LAYER, color=color)

    # Outer border
    box(x0, y0, W, H)

    # Horizontal dividers
    for frac in [0.6, 0.35, 0.15]:
        y = y0 + H*frac
        msp.add_line((x0, y), (x0+W, y),
                     dxfattribs={"layer": TITLE_LAYER, "color": 8})

    # Vertical dividers (3 columns)
    col1 = x0 + W*0.38
    col2 = x0 + W*0.65
    for x in [col1, col2]:
        msp.add_line((x, y0), (x, y0+H),
                     dxfattribs={"layer": TITLE_LAYER, "color": 8})

    # Row 1: company name (large)
    label("MAX FRANKE.led", x0+PAD, y0+H*0.65+PAD, h=TH*1.5, color=COLOR_A)
    label("Osdorfer Landstrasse 174-176  ·  D-22549 Hamburg",
          x0+PAD, y0+H*0.65-LH*0.5, h=LH*0.9)
    label("info@max-franke.de  ·  www.max-franke.com",
          x0+PAD, y0+H*0.65-LH*2.0, h=LH*0.9)

    # Row 2: project data
    y2 = y0 + H*0.35
    label("Projekt:", x0+PAD, y2+H*0.18, h=LH)
    label(project_name, x0+PAD, y2+H*0.08, h=TH, color=7)
    label("Bauherr:", col1+PAD, y2+H*0.18, h=LH)
    label(customer, col1+PAD, y2+H*0.08, h=TH, color=7)
    label("Planinhalt:", col2+PAD, y2+H*0.18, h=LH)
    label(f"Deckenrasterplan — {concept_id}", col2+PAD, y2+H*0.08, h=TH, color=7)

    # Row 3: scale / date / totals
    y3 = y0 + H*0.15
    label("Maßstab:", x0+PAD, y3+H*0.10, h=LH)
    label(scale, x0+PAD, y3+H*0.02, h=TH)
    label("Datum:", col1+PAD, y3+H*0.10, h=LH)
    label(generated, col1+PAD, y3+H*0.02, h=TH)
    label("Leuchten gesamt:", col2+PAD, y3+H*0.10, h=LH)
    label(f"{total_lumi} Stk  ·  {total_w:.0f} W", col2+PAD, y3+H*0.02, h=TH)

    # Row 4: warning note
    label("Achtung: Alle Maße am Bau zu prüfen! · Attention: All dimensions to be checked locally!",
          x0+PAD, y0+PAD, h=LH*0.9, color=8)


# ─────────────────────────────────────────────────────────────────────────────
# Legend panel
# ─────────────────────────────────────────────────────────────────────────────

def _draw_legend(msp, x0: float, y0: float, result: PlacementResult):
    """
    Draw legend box showing each luminaire type with symbol, code, qty.
    """
    from collections import Counter
    ROWS     = []
    type_cnt = Counter(p.lumi_type for p in result.placed)
    seen_types = set()
    for p in result.placed:
        t = p.lumi_type
        if t not in seen_types:
            seen_types.add(t)
            ROWS.append((t, p.product_code, p.description,
                         p.wattage, int(p.beam_angle_deg),
                         type_cnt[t]))

    ROW_H = 16_000   # height per row in mm-units — 2× so icons are visually large
    W     = 120_000
    PAD   = 2_000
    TH    = 2_200
    LH    = 1_800
    H     = ROW_H * (len(ROWS) + 2) + 6_000

    def box(x, y, w, h, color=7):
        msp.add_lwpolyline([(x,y),(x+w,y),(x+w,y+h),(x,y+h)], close=True,
                           dxfattribs={"layer": LEGEND_LAYER, "color": color})

    # Outer box
    box(x0, y0, W, H)

    # Header
    msp.add_line((x0, y0+H-ROW_H), (x0+W, y0+H-ROW_H),
                 dxfattribs={"layer": LEGEND_LAYER, "color": 8})
    _txt(msp, "LEUCHTENLEGENDE / LEGEND",
         (x0+PAD, y0+H-ROW_H+PAD), TH, LEGEND_LAYER, color=7)

    # Cutout info row
    cut_y = y0 + H - ROW_H*2
    msp.add_line((x0, cut_y), (x0+W, cut_y),
                 dxfattribs={"layer": LEGEND_LAYER, "color": 8})
    _txt(msp, "Deckenausschnitt  AD:140 mm  EBT:110 mm  DA:128 mm",
         (x0+PAD, cut_y+PAD), LH, LEGEND_LAYER, color=8)

    # Rows
    # r_leg scaled so symbols match the visual weight of the text.
    # 0.44 × ROW_H = 3520 at ROW_H=8000.  Also use lineweight 50 (0.50 mm) on
    # every symbol entity so the outline is as bold as the filled text beside it.
    r_leg    = int(ROW_H * 0.44)   # 7 040 at ROW_H=16 000 — big, visible icons
    col_sym  = x0 + PAD
    col_code = x0 + r_leg*2 + PAD*4
    col_qty  = x0 + W - 20_000

    for i, (ltype, pcode, desc, watt, beam, qty) in enumerate(ROWS):
        ry  = y0 + H - ROW_H*(i+3) + PAD
        clr = LUMI_COLORS.get(ltype, 7)
        r   = r_leg   # use display radius throughout this row

        # Mini symbol — type-specific geometry, thick pen so outlines match text weight
        cx = col_sym + r
        cy = ry + r
        la = {"layer": LEGEND_LAYER, "color": clr, "lineweight": 50}
        if ltype == 'A':
            msp.add_circle((cx,cy), r,             dxfattribs=la)
            msp.add_circle((cx,cy), r*0.30,        dxfattribs=la)
            msp.add_line((cx-r*.6,cy),(cx+r*.6,cy),dxfattribs=la)
            msp.add_line((cx,cy-r*.6),(cx,cy+r*.6),dxfattribs=la)
        elif ltype == 'B':
            msp.add_lwpolyline([(cx-r,cy-r),(cx+r,cy-r),(cx+r,cy+r),(cx-r,cy+r)],
                               close=True, dxfattribs=la)
            msp.add_circle((cx,cy), r*0.20,        dxfattribs=la)
        elif ltype == 'C':
            msp.add_lwpolyline([(cx,cy-r),(cx+r,cy),(cx,cy+r),(cx-r,cy)],
                               close=True, dxfattribs=la)
            msp.add_circle((cx,cy), r*0.20,        dxfattribs=la)
        elif ltype == 'D':
            msp.add_lwpolyline([(cx-r*0.866,cy+r*0.5),(cx+r*0.866,cy+r*0.5),(cx,cy-r)],
                               close=True, dxfattribs=la)
            msp.add_circle((cx,cy), r*0.20,        dxfattribs=la)
        elif ltype == 'E':
            msp.add_circle((cx,cy), r,             dxfattribs=la)
            msp.add_line((cx-r*.6,cy-r*.6),(cx+r*.6,cy+r*.6),dxfattribs=la)
            msp.add_line((cx-r*.6,cy+r*.6),(cx+r*.6,cy-r*.6),dxfattribs=la)
        else:
            msp.add_circle((cx,cy), r,             dxfattribs=la)
            msp.add_circle((cx,cy), r*0.35,        dxfattribs=la)

        # Type letter — floated above the symbol centre so it doesn't overlap text columns
        _txt(msp, ltype, (cx-TH*0.25, cy+r*0.55), TH*0.85, LEGEND_LAYER, color=clr)

        # Product code + description — col_code already accounts for large symbol
        _txt(msp, f"{pcode}",           (col_code, ry+r+TH*0.2), LH*0.85, LEGEND_LAYER)
        _txt(msp, f"{desc}  {watt}W  {beam}°",
             (col_code, ry+r-LH*1.2), LH*0.8, LEGEND_LAYER, color=8)

        # Quantity
        _txt(msp, f"× {qty}", (col_qty, ry+r), TH, LEGEND_LAYER, color=clr)

        # Row divider
        msp.add_line((x0, ry-PAD), (x0+W, ry-PAD),
                     dxfattribs={"layer": LEGEND_LAYER, "color": 8})


# ─────────────────────────────────────────────────────────────────────────────
# Ceiling tile grid (tile boundary lines — lights sit at cell centres)
# ─────────────────────────────────────────────────────────────────────────────

def _draw_ceiling_grid(msp, classified: "ClassifiedPlan",
                       pitch_mm: float = 625.0) -> None:
    """
    Draw ceiling tile boundary lines on CEILING-GRID layer.

    Grid lines are drawn at `pitch_mm` intervals starting from the Startmaß
    origin (1000mm, 2000mm from the zone corner).  pitch_mm = 625mm is the
    Rossmann Referenzmaß Rasterdecke (tile module).  Luminaires sit at the
    A_center sub-position (312.5mm from tile corner) inside each tile.
    """
    sf_z = next((z for z in classified.zones if z.zone_type == 'sales_floor'), None)
    if sf_z is None:
        return
    b   = sf_z.polygon.bounds          # (xmin, ymin, xmax, ymax)
    ox  = b[0] + 1000.0                # Startmaß: light origin X
    oy  = b[1] + 2000.0                # Startmaß: light origin Y
    half = pitch_mm * 0.5              # 625mm — offset to tile boundary

    _grid_attr = {"layer": GRID_LAYER, "color": COLOR_GRID, "lineweight": 50}

    # Draw vertical lines (constant X) from ymin..ymax
    x = ox - half
    while x <= b[2] + half:
        if x >= b[0] - half:
            msp.add_line((x, b[1]), (x, b[3]), dxfattribs=_grid_attr)
        x += pitch_mm

    # Draw horizontal lines (constant Y) from xmin..xmax
    y = oy - half
    while y <= b[3] + half:
        if y >= b[1] - half:
            msp.add_line((b[0], y), (b[2], y), dxfattribs=_grid_attr)
        y += pitch_mm


# ─────────────────────────────────────────────────────────────────────────────
# Dimension chains (showing 625mm tile grid and 1250mm inter-luminaire spacing)
# ─────────────────────────────────────────────────────────────────────────────

def _draw_dimensions(msp, result: PlacementResult, pitch_mm: float = 625.0):
    """
    Add a row and column of linear dimension annotations showing the grid pitch.
    Uses DXF DIMENSION entities with the standard dimension style.
    """
    if not result.placed:
        return

    # Find one row of type-A luminaires with consecutive X positions
    a_pts = sorted([(p.x, p.y) for p in result.placed if p.lumi_type == 'A'],
                   key=lambda pt: (round(pt[1]/pitch_mm), pt[0]))

    if len(a_pts) < 3:
        return

    # Pick the longest horizontal run
    by_row = defaultdict(list)
    for x, y in a_pts:
        by_row[round(y/pitch_mm)].append(x)
    longest_row_key = max(by_row, key=lambda k: len(by_row[k]))
    row_y = longest_row_key * pitch_mm
    row_xs = sorted(by_row[longest_row_key])[:5]  # show first 5 spacings

    dim_y = row_y + pitch_mm * 1.5   # offset above the row

    for i in range(len(row_xs)-1):
        x1, x2 = row_xs[i], row_xs[i+1]
        spacing = x2 - x1
        if abs(spacing - pitch_mm) < pitch_mm * 0.3:  # only annotate regular spacings
            # Leader lines down to luminaires
            msp.add_line((x1, row_y), (x1, dim_y),
                         dxfattribs={"layer": DIM_LAYER, "color": COLOR_DIM})
            msp.add_line((x2, row_y), (x2, dim_y),
                         dxfattribs={"layer": DIM_LAYER, "color": COLOR_DIM})
            # Horizontal dimension line
            msp.add_line((x1, dim_y), (x2, dim_y),
                         dxfattribs={"layer": DIM_LAYER, "color": COLOR_DIM})
            # Tick marks
            tk = pitch_mm * 0.05
            for xp in (x1, x2):
                msp.add_line((xp-tk, dim_y-tk), (xp+tk, dim_y+tk),
                             dxfattribs={"layer": DIM_LAYER, "color": COLOR_DIM})
            # Text
            _txt(msp, f"{spacing:.0f}", ((x1+x2)/2, dim_y+pitch_mm*0.1),
                 pitch_mm*0.12, DIM_LAYER, color=COLOR_DIM,
                 align=TextEntityAlignment.MIDDLE_CENTER)

    # "Startmaß Rasterdecke" annotation
    if row_xs:
        _txt(msp, "▶ Startmaß Rasterdecke",
             (row_xs[0] - pitch_mm*0.5, row_y - pitch_mm*2),
             pitch_mm * 0.12, DIM_LAYER, color=2)


# ─────────────────────────────────────────────────────────────────────────────
# M6  DXF export
# ─────────────────────────────────────────────────────────────────────────────

def export_dwg(result: PlacementResult, classified: ClassifiedPlan,
               source_dxf_path: Optional[str] = None,
               output_path: Optional[str] = None,
               project_name: str = "Lighting Project",
               customer: str = "Dirk Rossmann GmbH",
               concept_id: str = "rossmann_standard",
               scale: str = "1:75") -> Path:

    # Start from the source DXF if provided (preserves floor plan geometry)
    if source_dxf_path and Path(source_dxf_path).exists():
        try:
            doc = ezdxf.readfile(source_dxf_path)
        except Exception:
            doc = ezdxf.new("R2018")
    else:
        doc = ezdxf.new("R2018")

    doc.header["$INSUNITS"] = 4   # 4 = millimetres

    msp = doc.modelspace()

    # ── Ensure layers ─────────────────────────────────────────────────────
    for name, color, lw in [
        (LUMI_LAYER,   COLOR_A, 35),
        (ZONE_LAYER,   COLOR_ZONE, 18),
        (GRID_LAYER,   COLOR_GRID, 50),
        (DIM_LAYER,    COLOR_DIM, 18),
        (TITLE_LAYER,  7, 25),
        (LEGEND_LAYER, 7, 18),
        (ANNO_LAYER,   7, 13),
    ]:
        _add_layer(doc, name, color, lw)

    # ── Zone outlines ─────────────────────────────────────────────────────
    for zone in classified.zones:
        b = zone.polygon.bounds
        msp.add_lwpolyline(
            [(b[0],b[1]),(b[2],b[1]),(b[2],b[3]),(b[0],b[3])],
            close=True,
            dxfattribs={"layer": ZONE_LAYER, "color": COLOR_ZONE})
        _txt(msp, f"{zone.zone_type}  {zone.area_m2:.1f}m²",
             ((b[0]+b[2])/2, (b[1]+b[3])/2),
             800, ZONE_LAYER, color=COLOR_ZONE,
             align=TextEntityAlignment.MIDDLE_CENTER)

    # ── Luminaire blocks ──────────────────────────────────────────────────
    type_seen: dict[str, str] = {}  # lumi_type → block_name
    for lp in result.placed:
        color = LUMI_COLORS.get(lp.lumi_type, COLOR_A)
        if lp.lumi_type not in type_seen:
            bn = _ensure_lumi_block(doc, lp.lumi_type, lp.product_code,
                                    lp.cutout_mm, color)
            type_seen[lp.lumi_type] = bn
        else:
            bn = type_seen[lp.lumi_type]

        ref = msp.add_blockref(bn, insert=(lp.x, lp.y), dxfattribs={
            "layer": LUMI_LAYER, "rotation": lp.rotation, "color": color})

        # Attach ATTRIB values to the INSERT
        ref.add_attrib("TYPE",    lp.lumi_type,    insert=(lp.x, lp.y + lp.cutout_mm))
        ref.add_attrib("PRODUCT", lp.product_code, insert=(lp.x, lp.y - lp.cutout_mm))

    # ── Ceiling tile grid (tile boundaries — lights at cell centres) ──────
    _draw_ceiling_grid(msp, classified, pitch_mm=625.0)

    # ── Dimension annotations ──────────────────────────────────────────────
    _draw_dimensions(msp, result, pitch_mm=625.0)

    # ── Legend (top-right of drawing) ─────────────────────────────────────
    if classified.zones:
        b_all = classified.zones[0].polygon.bounds
        for z in classified.zones:
            b = z.polygon.bounds
            b_all = (min(b_all[0],b[0]), min(b_all[1],b[1]),
                     max(b_all[2],b[2]), max(b_all[3],b[3]))
        legend_x = b_all[2] + 10_000
        legend_y = b_all[3] - 80_000
        _draw_legend(msp, legend_x, legend_y, result)

        # ── Title block (below the drawing) ──────────────────────────────
        title_x = b_all[0]
        title_y = b_all[1] - 80_000
        _draw_title_block(
            msp, title_x, title_y,
            project_name=project_name,
            customer=customer,
            concept_id=concept_id,
            scale=scale,
            total_lumi=len(result.placed),
            total_w=result.total_wattage(),
            generated=datetime.datetime.now().strftime("%d.%m.%Y %H:%M"),
        )

    if output_path is None:
        output_path = str(EXPORTS_DIR /
                          f"{Path(result.source_file).stem}_luminaires.dxf")
    doc.saveas(output_path)
    print(f"DXF → {output_path}")
    return Path(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# M7  Excel BOM
# ─────────────────────────────────────────────────────────────────────────────
BLUE = PatternFill("solid", fgColor="1F3864")
ALT  = PatternFill("solid", fgColor="EBF0FA")
HF   = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
BF   = Font(size=9,  name="Calibri")
TF   = Font(bold=True, size=13, name="Calibri", color="1F3864")
TH   = Side(style="thin", color="BFBFBF")
CB   = Border(left=TH, right=TH, top=TH, bottom=TH)
CC   = Alignment(horizontal="center",  vertical="center", wrap_text=True)
LC   = Alignment(horizontal="left",    vertical="center", wrap_text=True)


def _hc(ws, row, col, val, w=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = HF; c.fill = BLUE; c.alignment = CC; c.border = CB
    if w: ws.column_dimensions[get_column_letter(col)].width = w


def _dc(ws, row, col, val, alt=False, align=CC):
    c = ws.cell(row=row, column=col, value=val)
    c.font = BF; c.alignment = align; c.border = CB
    if alt: c.fill = ALT


def export_excel(result: PlacementResult, classified: ClassifiedPlan,
                 project_name: str = "Lighting Project",
                 customer: str = "Dirk Rossmann GmbH",
                 concept_id: str = "rossmann_standard",
                 output_path: Optional[str] = None) -> Path:
    wb = openpyxl.Workbook()

    # ── Cover ────────────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Cover"; ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 36
    ws['A1'] = project_name;  ws['A1'].font = TF
    ws['A2'] = f"Customer: {customer}"
    ws['A2'].font = Font(size=10, color="444444", name="Calibri")
    ws['A3'] = f"Concept: {concept_id}"
    ws['A3'].font = Font(size=10, color="444444", name="Calibri")
    ws['A4'] = f"Generated: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws['A4'].font = Font(size=9, color="888888", name="Calibri", italic=True)

    summary_rows = [
        ("Total luminaires",               len(result.placed)),
        ("A — K1 Regalbeleuchtung 15W 40°",         len(result.by_type("A"))),
        ("AW — K1 Außenwand Beam-M-high 20W 40°",  len(result.by_type("AW"))),
        ("B — K4 Ergänzung 20W 60°",               len(result.by_type("B"))),
        ("C — K3 Rand 15W 40°",                    len(result.by_type("C"))),
        ("D — K2 Checkout/Service 20W 40°",        len(result.by_type("D"))),
        ("E — K6 Schaufenster 20W 60°",            len(result.by_type("E"))),
        ("W — Wabeneinsatz 20W 36°",               len(result.by_type("W"))),
        ("P — K5 Plakate 16W 24°",                 len(result.by_type("P"))),
        ("Total connected load",                   f"{result.total_wattage():.0f} W"),
        ("W/m² density",                           f"{result.total_wattage()/max(sum(z.area_m2 for z in classified.zones),1):.2f} W/m²"),
        ("Zones classified",                       len(classified.zones)),
        ("Tile module (Referenzmaß Rasterdecke)",  "625 mm"),
        ("Inter-luminaire spacing",                "1250 mm (2 tiles)"),
        ("Ceiling height (UK Rasterdecke)", "3000 mm"),
        ("Product family",                  "MIKA80-E + NEO85-SX (MAX FRANKE.led)"),
        ("Standard",                        "EN 12464-1 Lichtstrommethode"),
        ("Maintenance factor MF",           "0.80"),
    ]
    for i, (lbl, val) in enumerate(summary_rows, start=6):
        ws.cell(row=i, column=1, value=lbl).font = Font(bold=True, size=10, name="Calibri")
        ws.cell(row=i, column=2, value=val).font  = Font(size=10, name="Calibri")

    # ── Fixture BOM ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Fixture BOM"); ws2.freeze_panes = "A3"
    hdrs = [("Zone",18),("Type",7),("Product code",44),("Description",36),
            ("Mfr.",14),("W",8),("lm",8),("Beam°",8),("Qty",7),
            ("Total W",9),("Mounting",16),("IP",7),("Dim.",9)]
    for col, (h, w) in enumerate(hdrs, 1):
        _hc(ws2, 2, col, h, w)
    ws2.row_dimensions[2].height = 30

    agg = defaultdict(lambda: {"qty": 0, "lp": None})
    for lp in result.placed:
        k = (lp.zone_type, lp.lumi_type, lp.product_code)
        agg[k]["qty"] += 1; agg[k]["lp"] = lp

    row = 3; tot_qty = 0; tot_w = 0
    for (zt, lt, pc), d in sorted(agg.items()):
        lp = d["lp"]; q = d["qty"]; alt = (row % 2 == 0)
        vals = [zt.replace('_',' ').title(), lt, pc, lp.description, lp.manufacturer,
                lp.wattage, lp.lux_output, int(lp.beam_angle_deg), q, q*lp.wattage,
                lp.mounting_type.replace('_',' '), lp.ip_rating,
                "Yes" if lp.dimmable else "No"]
        for col, v in enumerate(vals, 1):
            _dc(ws2, row, col, v, alt=alt, align=LC if col in (1,3,4,5,11) else CC)
        tot_qty += q; tot_w += q * lp.wattage; row += 1

    for col in range(1, len(hdrs)+1):
        c = ws2.cell(row=row, column=col); c.border = CB; c.fill = BLUE; c.font = HF
    ws2.cell(row=row, column=4, value="TOTAL").font = HF
    ws2.cell(row=row, column=4).fill = BLUE
    ws2.cell(row=row, column=9, value=tot_qty).font = HF
    ws2.cell(row=row, column=9).fill = BLUE
    ws2.cell(row=row, column=10, value=tot_w).font = HF
    ws2.cell(row=row, column=10).fill = BLUE

    # ── Full schedule ────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Full Schedule"); ws3.freeze_panes = "A2"
    fh = [("#",5),("Zone",16),("Type",6),("Product code",44),("Description",32),
          ("W",6),("lm",7),("Beam°",7),("X mm",10),("Y mm",10),("Mounting",16),
          ("Grid",7),("Shelf",7),("IP",7),("CRI",6),("CCT K",7)]
    for col, (h, w) in enumerate(fh, 1):
        _hc(ws3, 1, col, h, w)
    ws3.row_dimensions[1].height = 30
    for idx, lp in enumerate(result.placed, 1):
        alt = (idx % 2 == 0)
        vals = [idx, lp.zone_type.replace('_',' ').title(), lp.lumi_type,
                lp.product_code, lp.description, lp.wattage, lp.lux_output,
                int(lp.beam_angle_deg), round(lp.x), round(lp.y),
                lp.mounting_type.replace('_',' '),
                "✓" if lp.grid_snapped else "–",
                "✓" if lp.shelf_aligned else "–",
                lp.ip_rating, lp.cri, lp.cct_k]
        for col, v in enumerate(vals, 1):
            _dc(ws3, idx+1, col, v, alt=alt,
                align=LC if col in (2,4,5,11) else CC)

    # ── Beleuchtungsberechnung (EN 12464-1 lighting calculation) ─────────────
    if result.zone_reports:
        ws4 = wb.create_sheet("Beleuchtungsberechnung"); ws4.freeze_panes = "A3"
        ws4.sheet_view.showGridLines = False
        ws4.column_dimensions['A'].width = 18

        # Title row
        t = ws4.cell(row=1, column=1, value="Beleuchtungsberechnung — EN 12464-1 Lichtstrommethode")
        t.font = TF; t.alignment = LC

        lh2 = [
            ("Zone",           16), ("Fläche m²",   10), ("Breite m",    9),
            ("Tiefe m",         9), ("Raumhöhe m",  10), ("Raumindex k",  10),
            ("Nutzungsgrad η",  12), ("Ziel-Em Lux", 12), ("n berechnet",  11),
            ("n platziert",    11), ("Raster mm",   10), ("Em erreicht",  11),
            ("Status",          8),
        ]
        for col, (h, w) in enumerate(lh2, 1):
            _hc(ws4, 2, col, h, w)
        ws4.row_dimensions[2].height = 30

        for row_i, rpt in enumerate(result.zone_reports, start=3):
            em_actual = rpt.maintained_lux_actual()
            ok        = "✓" if em_actual >= rpt.target_lux * 0.80 else "!"
            alt       = (row_i % 2 == 0)
            vals = [
                rpt.zone_type.replace('_', ' ').title(),
                round(rpt.area_m2, 1),
                rpt.room_width_m,
                rpt.room_depth_m,
                rpt.ceiling_height_m,
                round(rpt.room_index_k, 2),
                round(rpt.utilisation_factor, 3),
                rpt.target_lux,
                rpt.required_count,
                rpt.placed_count,
                rpt.grid_pitch_mm,
                round(em_actual, 0),
                ok,
            ]
            for col, v in enumerate(vals, 1):
                align = LC if col == 1 else CC
                c = ws4.cell(row=row_i, column=col, value=v)
                c.font = BF; c.alignment = align; c.border = CB
                if alt: c.fill = ALT
                # Colour-code status cell
                if col == 13:
                    c.font = Font(bold=True, size=9, name="Calibri",
                                  color="00AA00" if ok == "✓" else "CC0000")

        # Footer note
        note_row = 3 + len(result.zone_reports) + 1
        n = ws4.cell(row=note_row, column=1,
                     value="Wartungsfaktor MF = 0.80 (vierteljährl. Reinigung, LED) · Arbeitsebenenhöhe 850 mm")
        n.font = Font(size=8, italic=True, color="666666", name="Calibri")

    if output_path is None:
        output_path = str(EXPORTS_DIR /
                          f"{Path(result.source_file).stem}_fixture_schedule.xlsx")
    wb.save(output_path)
    print(f"Excel → {output_path}")
    return Path(output_path)


# ─────────────────────────────────────────────────────────────────────────────
# M8  PDF / HTML documentation
# ─────────────────────────────────────────────────────────────────────────────

_TPL = """<!DOCTYPE html><html><head><meta charset="utf-8"><style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Arial,sans-serif;font-size:9pt;color:#1a1a2e}
.cov{background:#1F3864;color:#fff;padding:44px}
.cov h1{font-size:18pt;font-weight:700;margin-bottom:6px}
.cov .s{font-size:9pt;opacity:.8;margin-top:4px}
.cov .m{font-size:7.5pt;opacity:.55;margin-top:10px}
.sec{padding:18px 44px}
h2{font-size:10pt;font-weight:700;color:#1F3864;border-bottom:2px solid #1F3864;
   padding-bottom:3px;margin-bottom:8px;margin-top:14px}
.row{display:flex;gap:10px;margin:10px 0}
.st{background:#EBF0FA;border-left:4px solid #1F3864;padding:8px 12px;flex:1}
.st .v{font-size:14pt;font-weight:700;color:#1F3864}
.st .l{font-size:7pt;color:#555;margin-top:1px}
table{width:100%;border-collapse:collapse;font-size:8pt;margin-top:6px}
th{background:#1F3864;color:#fff;padding:5px 7px;text-align:left}
td{padding:4px 7px;border-bottom:1px solid #dde4f0}
tr:nth-child(even) td{background:#f4f7fd}
.ft{background:#EBF0FA;padding:10px 44px;font-size:7pt;color:#666;
    border-top:1px solid #d0daea;margin-top:20px}
.badge{display:inline-block;padding:2px 6px;border-radius:3px;
       font-size:7pt;font-weight:700;margin-left:4px}
.badge-a{background:#e040fb22;color:#e040fb}
.badge-b{background:#f4433622;color:#f44336}
.badge-c{background:#00bcd422;color:#00bcd4}
.badge-d{background:#ffeb3b22;color:#f9a825}
.badge-e{background:#2196f322;color:#1565c0}
.badge-w{background:#4caf5022;color:#2e7d32}
.badge-p{background:#ff980022;color:#e65100}
@page{size:A4;margin:0}
</style></head><body>
<div class="cov">
  <h1>{{ project_name }}</h1>
  <div class="s">Deckenrasterplan / Lighting Design Documentation — {{ concept_id }}</div>
  <div class="s">{{ customer }}</div>
  <div class="m">Generated {{ generated }} · lighting-ai · MAX FRANKE.led MIKA80-E</div>
</div>
<div class="sec"><h2>Summary</h2><div class="row">
  <div class="st"><div class="v">{{ total }}</div><div class="l">Total luminaires</div></div>
  <div class="st"><div class="v">{{ tw }} W</div><div class="l">Connected load</div></div>
  <div class="st"><div class="v">
    <span class="badge badge-a">{{ ta }}×A</span>
    <span class="badge badge-b">{{ tb }}×B</span>
    {% if tc %}<span class="badge badge-c">{{ tc }}×C</span>{% endif %}
    {% if td %}<span class="badge badge-d">{{ td }}×D</span>{% endif %}
    {% if te %}<span class="badge badge-e">{{ te }}×E</span>{% endif %}
    {% if tw_type %}<span class="badge badge-w">{{ tw_type }}×W</span>{% endif %}
    {% if tp %}<span class="badge badge-p">{{ tp }}×P</span>{% endif %}
  </div><div class="l">A=K1 15W 40° / B=K4 20W 60° / C=K3 Rand / D=K2 Checkout / E=K6 Track / W=Wabe / P=Plakate</div></div>
  <div class="st"><div class="v">{{ zones }}</div><div class="l">Zones classified</div></div>
</div></div>
<div class="sec"><h2>Zone Summary</h2><table>
<tr><th>Zone</th><th>Type</th><th>Area m²</th><th>Qty</th>
    <th>Product</th><th>Method</th><th>Confidence</th></tr>
{% for r in zr %}<tr>
  <td>{{ r.label }}</td><td>{{ r.zt }}</td><td>{{ r.area }}</td>
  <td>{{ r.qty }}</td><td style="font-size:7pt">{{ r.prod }}</td>
  <td>{{ r.method }}</td><td>{{ r.conf }}%</td>
</tr>{% endfor %}
</table></div>
<div class="sec"><h2>Fixture Schedule (Summary)</h2><table>
<tr><th>Type</th><th>Product code</th><th>Description</th>
    <th>Qty</th><th>W</th><th>Total W</th><th>Beam</th><th>IP</th></tr>
{% for r in br %}<tr>
  <td>{{ r.t }}</td><td style="font-size:7pt">{{ r.c }}</td>
  <td>{{ r.d }}</td><td>{{ r.q }}</td><td>{{ r.w }}</td>
  <td><b>{{ r.tw }}</b></td><td>{{ r.b }}°</td><td>{{ r.ip }}</td>
</tr>{% endfor %}
<tr style="background:#1F3864;color:#fff;font-weight:700">
  <td colspan="3">TOTAL</td><td>{{ total }}</td><td></td>
  <td>{{ tw }} W</td><td></td><td></td>
</tr>
</table></div>
<div class="sec"><h2>Technical Specs</h2><table>
{% for k,v in specs %}<tr><td><b>{{ k }}</b></td><td>{{ v }}</td></tr>{% endfor %}
</table></div>
<div class="ft">
  Auto-generated by lighting-ai v1.0 · Subject to designer review ·
  Alle Maße am Bau prüfen! · MAX FRANKE.led · info@max-franke.de
</div>
</body></html>"""


def export_pdf(result: PlacementResult, classified: ClassifiedPlan,
               concept_id: str = "rossmann_standard",
               customer: str = "Dirk Rossmann GmbH",
               project_name: str = "Lighting Project",
               output_path: Optional[str] = None) -> Path:
    bom = defaultdict(lambda: {"qty": 0, "lp": None})
    for lp in result.placed:
        bom[lp.product_code]["qty"] += 1; bom[lp.product_code]["lp"] = lp

    br = [{"t": d["lp"].lumi_type, "c": k, "d": d["lp"].description,
           "q": d["qty"], "w": d["lp"].wattage,
           "tw": d["qty"]*d["lp"].wattage,
           "b": int(d["lp"].beam_angle_deg),
           "ip": d["lp"].ip_rating}
          for k, d in sorted(bom.items())]

    zc = Counter(lp.zone_type for lp in result.placed)
    zp = {lp.zone_type: lp.product_code for lp in result.placed}
    zr = [{"label": f"Zone {z.polygon_index}",
           "zt":   z.zone_type.replace('_',' ').title(),
           "area": round(z.area_m2, 1),
           "qty":  zc.get(z.zone_type, 0),
           "prod": zp.get(z.zone_type, '—'),
           "method": z.method,
           "conf": round(z.confidence*100)}
          for z in classified.zones]

    specs = [
        ("Tile module (Referenzmaß Rasterdecke)", "625 mm"),
        ("Inter-luminaire spacing (shelf runs)",  "1250 mm (2 tiles)"),
        ("Ceiling height (UK Rasterdecke)",       "3000 mm / Fries 3300 mm"),
        ("A — K1 Regalbeleuchtung Innenraum",    "MIKA80-E 15W 40° 2400lm"),
        ("AW — K1 Außenwand Beam-M-high",        "MIKA80-E 20W 40° 3200lm"),
        ("B — K4 Ergänzungsbeleuchtung",          "MIKA80-E 20W 60° 3200lm"),
        ("C — K3 Regalbeleuchtung Rand",          "MIKA80-E 15W 40° 2400lm"),
        ("D — K2 Checkout / Service",             "MIKA80-E 20W 40° 3200lm"),
        ("E — K6 Schaufenster",                   "NEO85-SX 20W 60° 3200lm Track"),
        ("W — Wabeneinsatz Cosmetics",            "MIKA80-E 20W 36° 1700lm Honeycomb"),
        ("P — K5 Plakate Poster",                 "MIKA80-E 16W 24° 2100lm Power-Linse"),
        ("Cutout dia. (DA)",                   "128 mm  /  NEO85-SX: 85 mm"),
        ("Outer dia. (AD)",                    "140 mm  /  NEO85-SX: 85 mm"),
        ("Embed depth (EBT)",                  "110 mm  /  NEO85-SX: 146 mm"),
        ("CCT",                                "3000 K"),
        ("CRI",                                ">90"),
        ("Dimmable",                           "Yes (DV2.5)"),
        ("IP rating (all)",                    "IP20"),
        ("Tilt / Rotate",                      "35° / 355°"),
        ("Maintenance factor MF",              "0.80  (CIE 97, quarterly cleaning, LED)"),
        ("Work plane",                         "850 mm  (EN 12464-1)"),
        ("No-lighting zones",                  "Windfang, Rolltreppe, Aufzug, WC, Technik"),
    ]
    html = Template(_TPL).render(
        project_name=project_name, concept_id=concept_id, customer=customer,
        generated=datetime.datetime.now().strftime("%d.%m.%Y %H:%M"),
        total=len(result.placed),
        tw=round(result.total_wattage()),
        ta=len(result.by_type("A")),
        tb=len(result.by_type("B")),
        tc=len(result.by_type("C")),
        td=len(result.by_type("D")),
        te=len(result.by_type("E")),
        tw_type=len(result.by_type("W")),
        tp=len(result.by_type("P")),
        zones=len(classified.zones),
        br=br, zr=zr, specs=specs,
    )

    stem = Path(result.source_file).stem
    if output_path is None:
        output_path = str(EXPORTS_DIR / stem)

    # Try WeasyPrint first (proper PDF), fall back to HTML
    try:
        from weasyprint import HTML as WP
        out = output_path + ".pdf"
        WP(string=html).write_pdf(out)
        print(f"PDF → {out}")
        return Path(out)
    except Exception:
        pass

    out = output_path + ".html"
    Path(out).write_text(html, encoding="utf-8")
    print(f"HTML → {out}")
    return Path(out)


# ── CLI smoke test ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    from services.parser.pdf_parser import RealPlanParser
    from services.classifier.room_classifier_real import RealRoomClassifier
    from services.placer.real_placer import RealLuminairePlacer
    UP = Path("/mnt/user-data/uploads")
    plan       = RealPlanParser().parse(UP/"3600_HH_Jungfernstieg_EG_SB_Kassen_20240506.pdf")
    classified = RealRoomClassifier().classify(plan)
    result     = RealLuminairePlacer().place_all(plan, classified)
    print(result.summary())
    dwg  = export_dwg(result, classified,
                      project_name="Rossmann Hamburg Jungfernstieg EG",
                      customer="Dirk Rossmann GmbH")
    xlsx = export_excel(result, classified,
                        project_name="Rossmann Hamburg Jungfernstieg EG",
                        customer="Dirk Rossmann GmbH")
    pdf  = export_pdf(result, classified,
                      project_name="Rossmann Hamburg Jungfernstieg EG",
                      customer="Dirk Rossmann GmbH")
    print(f"\nDXF:  {dwg}\nXLSX: {xlsx}\nPDF:  {pdf}")
